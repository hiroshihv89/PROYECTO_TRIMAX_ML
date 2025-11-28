def train_model_retrasos(file_path, results_folder="results"):
    import os
    import joblib
    import matplotlib.pyplot as plt
    import seaborn as sns
    from sklearn.preprocessing import LabelEncoder
    from sklearn.ensemble import RandomForestClassifier
    from sklearn.model_selection import train_test_split
    from sklearn.metrics import classification_report, confusion_matrix, accuracy_score
    import pandas as pd
    import numpy as np

    try:
        # Cargar datos con validacion
        print(f"Cargando archivo: {file_path}")
        df = pd.read_excel(file_path)
        print(f"Datos cargados: {len(df)} registros, {len(df.columns)} columnas")
        print(f"Columnas disponibles: {list(df.columns)}")

        # Validacion de datos esenciales
        required_columns = ['FECHA_INICIO', 'FECHA_TERMINO', 'PLANTA', 'SEDE', 'TIPO', 'SIMTIPO', 'PRODUCTO', 'TIPO TRATAMIENTO']
        missing_columns = [col for col in required_columns if col not in df.columns]
        if missing_columns:
            raise ValueError(f"Columnas faltantes en el dataset: {missing_columns}")

        # Convertir fechas con manejo de errores
        df['FECHA_INICIO'] = pd.to_datetime(df['FECHA_INICIO'], errors='coerce')
        df['FECHA_TERMINO'] = pd.to_datetime(df['FECHA_TERMINO'], errors='coerce')
        fecha_original_count = len(df)
        df = df.dropna(subset=['FECHA_INICIO', 'FECHA_TERMINO'])
        fecha_filtrada_count = len(df)
        
        if fecha_original_count != fecha_filtrada_count:
            print(f"Se eliminaron {fecha_original_count - fecha_filtrada_count} registros con fechas invalidas")

        # Variables temporales enriquecidas (basadas en FECHA_INICIO)
        df['anio'] = df['FECHA_INICIO'].dt.year
        df['mes'] = df['FECHA_INICIO'].dt.month
        df['dia'] = df['FECHA_INICIO'].dt.day
        df['dia_semana'] = df['FECHA_INICIO'].dt.weekday
        df['semana_anual'] = df['FECHA_INICIO'].dt.isocalendar().week.astype(int)
        df['trimestre'] = df['FECHA_INICIO'].dt.quarter
        df['es_fin_de_semana'] = df['dia_semana'].isin([5, 6]).astype(int)
        df['hora_inicio'] = df['FECHA_INICIO'].dt.hour

        # Codificacion categorica mejorada
        categorical_cols = ['PLANTA', 'SEDE', 'TIPO', 'SIMTIPO', 'PRODUCTO', 'TIPO TRATAMIENTO']
        label_encoders = {}
        
        for col in categorical_cols:
            if col in df.columns:
                # Manejo de valores nulos
                df[col] = df[col].fillna("DESCONOCIDO")
                
                # Verificar si hay suficientes categorias
                unique_values = df[col].nunique()
                if unique_values == 1:
                    print(f"Columna '{col}' tiene solo una categoria, considerarla para eliminar")
                
                le = LabelEncoder()
                df[f'{col}_enc'] = le.fit_transform(df[col].astype(str))
                label_encoders[col] = le
                print(f"Codificada columna '{col}' con {unique_values} categorias")

        # Seleccion de caracteristicas mejorada
        features = [f'{col}_enc' for col in categorical_cols if col in df.columns] + \
                   ['anio', 'mes', 'dia', 'dia_semana', 'semana_anual', 'trimestre', 'es_fin_de_semana']

        # Validar que todas las features existen
        missing_features = [f for f in features if f not in df.columns]
        if missing_features:
            print(f"Features faltantes: {missing_features}")
            features = [f for f in features if f in df.columns]

        X = df[features].fillna(0)
        print(f"Features utilizadas: {len(features)}")
        print(f"Shape de X: {X.shape}")

        # Calcular la variable objetivo (retraso) desde las fechas
        if 'FECHA_INICIO' in df.columns and 'FECHA_TERMINO' in df.columns:
            print("Calculando retraso desde FECHA_INICIO y FECHA_TERMINO...")
            
            # Calcular cuantas horas tardo cada orden
            df['tiempo_proceso_horas'] = (df['FECHA_TERMINO'] - df['FECHA_INICIO']).dt.total_seconds() / 3600
            
            # Tiempos esperados segun el tipo de orden
            tiempos_esperados = {
                'FABRICACION': 24,
                'BISELADO': 8,
                'FABRICACION Y BISELADO': 32
            }
            
            def calcular_retraso(row):
                tipo = row['TIPO'] if pd.notna(row['TIPO']) else 'FABRICACION'
                tiempo_esperado = tiempos_esperados.get(tipo, 24)
                # Si tardo mas de lo esperado, es retraso
                return 1 if row['tiempo_proceso_horas'] > tiempo_esperado else 0
            
            df['retraso'] = df.apply(calcular_retraso, axis=1)
            print(f"Retrasos calculados: {df['retraso'].sum()} de {len(df)} ordenes ({df['retraso'].mean()*100:.2f}%)")
        elif 'retraso' not in df.columns:
            print("Columna 'retraso' no encontrada, creando variable objetivo dummy para prueba")
            df['retraso'] = np.random.randint(0, 2, size=len(df))
        
        # Analisis del balance de clases
        class_distribution = df['retraso'].value_counts(normalize=True)
        print(f"Distribucion de clases: {class_distribution.to_dict()}")
        
        y = df['retraso']

        # Entrenamiento con validacion mejorada
        X_train, X_test, y_train, y_test = train_test_split(
            X, y, 
            test_size=0.2, 
            random_state=42,
            stratify=y  # Mantener proporcion de clases
        )
        
        print(f"Datos de entrenamiento: {len(X_train)} registros")
        print(f"Datos de prueba: {len(X_test)} registros")

        # Configurar el modelo Random Forest
        model = RandomForestClassifier(
            n_estimators=200,
            max_depth=15,
            min_samples_split=15,
            min_samples_leaf=8,
            random_state=42,
            class_weight='balanced',
            n_jobs=-1  # Usar todos los cores disponibles
        )
        
        print("Entrenando modelo...")
        model.fit(X_train, y_train)

        # Evaluacion completa del modelo
        y_pred = model.predict(X_test)
        accuracy = accuracy_score(y_test, y_pred)
        
        print("Metricas de evaluacion:")
        print(classification_report(y_test, y_pred))
        
        # Guardar todos los resultados
        os.makedirs(results_folder, exist_ok=True)
        
        # Guardar el modelo entrenado y los encoders
        joblib.dump(model, os.path.join(results_folder, "modelo_retrasos.pkl"))
        joblib.dump(label_encoders, os.path.join(results_folder, "label_encoders.pkl"))
        
        # Agregar predicciones al dataset y guardarlo
        df['prediccion'] = model.predict(X)
        
        # Calcular probabilidad de retraso (manejar caso de una sola clase)
        probas = model.predict_proba(X)
        if probas.shape[1] == 2:
            df['probabilidad_retraso'] = probas[:, 1]
        else:
            df['probabilidad_retraso'] = probas[:, 0]
            print("Advertencia: Solo se detecto una clase en el modelo")
        
        # Agregar columna nivel_riesgo con 3 niveles
        def asignar_riesgo(prob):
            if prob < 0.4:
                return "Bajo"
            elif prob < 0.7:
                return "Medio"
            else:
                return "Alto"
        
        df['nivel_riesgo'] = df['probabilidad_retraso'].apply(asignar_riesgo)
        
        # Guardar Excel con colores
        excel_path = os.path.join(results_folder, "dataset_predicciones.xlsx")
        
        try:
            from openpyxl import load_workbook
            from openpyxl.styles import PatternFill
            
            # Guardar primero el Excel
            df.to_excel(excel_path, index=False, engine='openpyxl')
            
            # Aplicar colores a la columna nivel_riesgo
            wb = load_workbook(excel_path)
            ws = wb.active
            
            # Encontrar la columna nivel_riesgo
            riesgo_col = None
            for col in range(1, ws.max_column + 1):
                if ws.cell(row=1, column=col).value == 'nivel_riesgo':
                    riesgo_col = col
                    break
            
            if riesgo_col:
                # Colores de fondo en columna nivel_riesgo
                verde = PatternFill(start_color='90EE90', end_color='90EE90', fill_type='solid')
                amarillo = PatternFill(start_color='FFD700', end_color='FFD700', fill_type='solid')
                rojo = PatternFill(start_color='FF6B6B', end_color='FF6B6B', fill_type='solid')
                
                # Aplicar colores según el valor
                for row in range(2, ws.max_row + 1):
                    cell = ws.cell(row=row, column=riesgo_col)
                    if cell.value == 'Bajo':
                        cell.fill = verde
                    elif cell.value == 'Medio':
                        cell.fill = amarillo
                    elif cell.value == 'Alto':
                        cell.fill = rojo
                
                wb.save(excel_path)
                print(f"Excel guardado con colores en nivel_riesgo: {excel_path}")
            else:
                print("Advertencia: No se encontro la columna nivel_riesgo para aplicar colores")
        except Exception as e:
            print(f"Error aplicando colores al Excel: {str(e)}")
            # Si falla, guardar sin colores
            df.to_excel(excel_path, index=False)

        # Graficos mejorados
        plt.style.use('default')
        fig, axes = plt.subplots(2, 2, figsize=(15, 12))
        
        # 1. Top variables importantes
        importancias = pd.Series(model.feature_importances_, index=features).sort_values(ascending=False)
        if len(importancias) > 0:
            top_features = importancias.head(min(8, len(importancias)))
            sns.barplot(x=top_features.index, y=top_features.values, ax=axes[0,0])
            axes[0,0].set_title("Top Variables Importantes")
            axes[0,0].tick_params(axis='x', rotation=45)
        else:
            axes[0,0].text(0.5, 0.5, 'No hay features disponibles', ha='center', va='center')
            axes[0,0].set_title("Top Variables Importantes")
        
        # 2. Matriz de confusion
        cm = confusion_matrix(y_test, y_pred)
        sns.heatmap(cm, annot=True, fmt='d', cmap='Blues', ax=axes[0,1])
        axes[0,1].set_title("Matriz de Confusion")
        axes[0,1].set_xlabel("Predicho")
        axes[0,1].set_ylabel("Real")
        
        # 3. Retrasos por mes
        retrasos_por_mes = df.groupby('mes')['retraso'].mean() * 100
        if len(retrasos_por_mes) > 0:
            retrasos_por_mes.plot(kind='bar', ax=axes[1,0], color='skyblue')
            axes[1,0].set_title("Porcentaje de Retrasos por Mes")
            axes[1,0].set_ylabel("% Retrasos")
        else:
            axes[1,0].text(0.5, 0.5, 'No hay datos para el grafico', ha='center', va='center')
            axes[1,0].set_title("Porcentaje de Retrasos por Mes")
        
        # 4. Distribucion de probabilidades
        if 'probabilidad_retraso' in df.columns:
            df['probabilidad_retraso'].hist(bins=30, ax=axes[1,1], alpha=0.7)
            axes[1,1].set_title("Distribucion de Probabilidades de Retraso")
            axes[1,1].set_xlabel("Probabilidad")
            axes[1,1].set_ylabel("Frecuencia")
        else:
            axes[1,1].text(0.5, 0.5, 'No hay probabilidades calculadas', ha='center', va='center')
            axes[1,1].set_title("Distribucion de Probabilidades de Retraso")
        
        plt.tight_layout()
        plt.savefig(os.path.join(results_folder, "analisis_completo.png"), dpi=300, bbox_inches='tight')
        plt.close()

        # Analisis adicionales
        mes_max_retraso = df.groupby('mes')['retraso'].mean().idxmax() if len(df['mes'].unique()) > 0 else 0
        retrasos_por_mes_dict = (df.groupby('mes')['retraso'].mean() * 100).round(2).to_dict()
        
        # Top categorias con mas retrasos
        top_categorias_retraso = {}
        for col in categorical_cols:
            if col in df.columns and df[col].nunique() > 0:
                try:
                    retraso_por_categoria = df.groupby(col)['retraso'].mean().sort_values(ascending=False).head(3)
                    top_categorias_retraso[col] = retraso_por_categoria.to_dict()
                except Exception as e:
                    print(f"Error procesando categoria {col}: {e}")
                    top_categorias_retraso[col] = {}

        # Retornar resultados enriquecidos
        resultados = {
            "accuracy_test": float(accuracy),  # Asegurar que es float serializable
            "top_features": importancias[:10].to_dict(),
            "recomendacion_mes": int(mes_max_retraso),
            "retrasos_por_mes": retrasos_por_mes_dict,
            "distribucion_clases": class_distribution.to_dict(),
            "top_categorias_retraso": top_categorias_retraso,
            "registros_procesados": int(len(df)),
            "features_utilizadas": int(len(features))
        }
        
        # Guardar resumen de resultados
        with open(os.path.join(results_folder, "resumen_resultados.txt"), "w") as f:
            f.write("RESUMEN DE ENTRENAMIENTO DEL MODELO\n")
            f.write("===================================\n")
            f.write(f"Accuracy: {accuracy:.4f}\n")
            f.write(f"Registros procesados: {len(df)}\n")
            f.write(f"Features utilizadas: {len(features)}\n")
            f.write(f"Mes con mas retrasos: {mes_max_retraso}\n")
            f.write(f"Distribucion de clases: {class_distribution.to_dict()}\n")
        
        print("Entrenamiento completado exitosamente")
        print(f"Resultados a retornar: {resultados}")
        
        return resultados

    except Exception as e:
        print(f"ERROR CRITICO en train_model_retrasos: {str(e)}")
        print(f"Tipo de error: {type(e).__name__}")
        import traceback
        print(f"Traceback completo:\n{traceback.format_exc()}")
        raise  # Re-lanzar la excepcion para que el backend la capture
